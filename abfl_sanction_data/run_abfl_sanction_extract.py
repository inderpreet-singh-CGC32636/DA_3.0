"""
ABFL Sanction Format — 132-Column Data Extraction
==================================================
Pulls data from Redshift for all 297 HE loans and writes to an
Excel file matching the "Required for pool shortlisting" sheet in
Sancttion_Format.xlsb.

Columns that CANNOT be sourced from the DB are left blank and
highlighted ORANGE in the output header row.

Run:
    python ABFL_Formats/run_abfl_sanction_extract.py
"""
from __future__ import annotations

import logging
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── paths ───────────────────────────────────────────────────────────────────
ROOT         = Path(__file__).resolve().parents[1]
ABFL_DIR     = ROOT / "ABFL_Formats"
SOURCE_XLSM  = ABFL_DIR / "Pool Upload Format - Final_.xlsm"   # LAN source
SQL_FILE     = ABFL_DIR / "sql_abfl_sanction_format.sql"
CONFIG_PATH  = ROOT / "config" / "database.yaml"
TIMESTAMP    = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_PATH  = ABFL_DIR / f"HE_sanction_format_{TIMESTAMP}.xlsx"
LOG_DIR      = ROOT / "logs"

SHEET_NAME   = "Required for pool shortlisting"

# ── header colours ───────────────────────────────────────────────────────────
FILL_GREEN   = PatternFill("solid", fgColor="C6EFCE")   # DB-filled column
FILL_ORANGE  = PatternFill("solid", fgColor="FF8C00")   # Not found in DB
FILL_BLUE    = PatternFill("solid", fgColor="BDD7EE")   # Hardcoded/derived
FONT_DARK    = Font(color="1F3864", bold=True)
FONT_WHITE   = Font(color="FFFFFF", bold=True)

# Columns intentionally returned as NULL (no DB source) → ORANGE header
NOT_FOUND = {
    "Subvention Scheme if any",
    "PMAY subsidy status \n(Claimed & received/ Claimed & not received / Not claimed/Claimed but rejected)",
    "NRI Loan (Yes/No)",
    "ECLGS",
    "Link loan Part of pool (Yes/ No)",
    "Link Loan/Top up loan Number if applicable",
    "Legal Report ",
}

# Hardcoded / derived columns (not NULL, but not raw DB field) → BLUE header
HARDCODED = {
    "PMAY Flag (Y/N) - All are HE cases thus marked as No",
    "EMI Frequency",
    "Customer Type (Primary Applicant)",
    "Customer Type (Applicant 2)",
    "Customer Type (Applicant 3)",
    "Customer Type (Applicant 4)",
    "Customer Type (Applicant 5)",
    "Customer Type (Applicant 6)",
    "Customer Type (Applicant 7)",
    "Customer Type (Applicant 8)",
    "Customer Type (Applicant 9)",
    "Open Plot (Y/N)",
    "Current LTV at asset level",
    "fully disbursed Yes/No",
    "Staff Loan (Yes/No)",
    "Restructured Flag (including OTR 1/OTR 2)",
    "Loan became NPA since origination (Y/N)",
}


# ── logging ──────────────────────────────────────────────────────────────────
def setup_logging() -> logging.Logger:
    LOG_DIR.mkdir(exist_ok=True)
    fmt = "%(asctime)s %(levelname)s %(message)s"
    logging.basicConfig(
        level=logging.INFO,
        format=fmt,
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(LOG_DIR / f"sanction_extract_{TIMESTAMP}.log", encoding="utf-8"),
        ],
    )
    return logging.getLogger(__name__)


# ── DB helpers ───────────────────────────────────────────────────────────────
def load_db_config() -> dict:
    raw = yaml.safe_load(CONFIG_PATH.read_text())
    return raw.get("database", raw)


def get_connection(cfg: dict):
    import redshift_connector
    return redshift_connector.connect(
        host=cfg["host"],
        port=int(cfg.get("port", 5439)),
        database=cfg["database"],
        user=cfg["user"],
        password=cfg["password"],
        timeout=300,
    )


# ── LAN list ─────────────────────────────────────────────────────────────────
def get_lan_list(logger: logging.Logger) -> str:
    """Read sz_loan_account_no values from the Deal sheet of the source xlsm."""
    logger.info("Reading LAN list from: %s", SOURCE_XLSM.name)
    raw = pd.read_excel(SOURCE_XLSM, sheet_name="Deal and Loan data",
                        header=None, engine="openpyxl")
    # find header row (first row with most non-null values)
    best_row, best_cnt = 0, 0
    for i, row in raw.iterrows():
        cnt = row.notna().sum()
        if cnt > best_cnt:
            best_cnt, best_row = cnt, i
        if i > 10:
            break
    df = pd.read_excel(SOURCE_XLSM, sheet_name="Deal and Loan data",
                       header=best_row, engine="openpyxl")
    # find the application/loan number column
    loan_col = None
    for c in df.columns:
        if "application form" in str(c).lower() or "loan no" in str(c).lower():
            loan_col = c
            break
    if loan_col is None:
        loan_col = df.columns[0]

    lans = df[loan_col].dropna().astype(str).str.strip()
    # keep only numeric-looking LANs
    lans = lans[lans.str.match(r"^\d+$")]
    lans = lans[lans.str.len() > 5].unique().tolist()
    logger.info("  %d LANs found", len(lans))
    return "'" + "','".join(lans) + "'"


# ── SQL parsing ──────────────────────────────────────────────────────────────
def load_sql(lan_list: str) -> str:
    """Read the SQL file, strip trailing comment block, inject LAN list."""
    raw = SQL_FILE.read_text(encoding="utf-8")
    # drop the trailing comment block (starts after the last semicolon section)
    parts = raw.split("\n\n\n-- ===")
    sql_body = parts[0].strip()
    return sql_body.replace("{LAN_LIST}", lan_list)


# ── query runner ─────────────────────────────────────────────────────────────
def run_query(conn, sql: str, logger: logging.Logger) -> pd.DataFrame:
    logger.info("Executing sanction format SQL (%d chars) ...", len(sql))
    cur = conn.cursor()
    cur.execute(sql)
    cols = [d[0] for d in cur.description]
    rows = cur.fetchall()
    cur.close()
    df = pd.DataFrame(rows, columns=cols)
    logger.info("  -> %d rows x %d columns", len(df), len(df.columns))
    return df


# ── Excel writer ─────────────────────────────────────────────────────────────
def write_excel(df: pd.DataFrame, logger: logging.Logger) -> None:
    logger.info("Writing output: %s", OUTPUT_PATH)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── header row ───────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border

        # choose highlight colour
        if col_name in NOT_FOUND:
            cell.fill = FILL_ORANGE
            cell.font = FONT_WHITE
        elif col_name in HARDCODED:
            cell.fill = FILL_BLUE
            cell.font = FONT_DARK
        else:
            cell.fill = FILL_GREEN
            cell.font = FONT_DARK

    # ── data rows ────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # ── column widths ────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        # header text length (first line only for wrapped headers)
        header_len = max(len(line) for line in str(col_name).splitlines())
        # sample data max length (up to first 50 rows)
        sample = df.iloc[:50, col_idx - 1].astype(str).str.len()
        max_val = sample.max()
        data_len = int(max_val) if (len(sample) > 0 and pd.notna(max_val)) else 0
        ws.column_dimensions[col_letter].width = min(max(header_len, data_len, 8), 40)

    # freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    logger.info("Saved: %s", OUTPUT_PATH)


# ── eval checks ──────────────────────────────────────────────────────────────
def eval_checks(df: pd.DataFrame, logger: logging.Logger) -> None:
    logger.info("=" * 70)
    logger.info("EVAL CHECKS")
    logger.info("=" * 70)
    logger.info("Total rows: %d  (expected 297)", len(df))

    # mandatory key cols
    key_cols = ["Loan No.", "Customer Name (Primary Applicant)",
                "Principal O/s", "Current ROI"]
    for c in key_cols:
        matches = [col for col in df.columns if col.lower() == c.lower()]
        if matches:
            nulls = int(df[matches[0]].isna().sum())
            logger.info("  Nulls in %-40s : %d", c, nulls)

    # duplicate LANs
    loan_col_matches = [c for c in df.columns if "loan no" in c.lower()]
    if loan_col_matches:
        dups = df[loan_col_matches[0]].duplicated().sum()
        logger.info("  Duplicate Loan No.                          : %d", dups)

    # fill rate summary (use iloc to avoid duplicate-column ambiguity)
    filled_cols  = sum(1 for i in range(len(df.columns)) if df.iloc[:, i].notna().any())
    orange_cols  = len(NOT_FOUND)
    logger.info("  Columns with data: %d / %d", filled_cols, len(df.columns))
    logger.info("  Columns not in DB (orange): %d", orange_cols)

    logger.info("=" * 70)
    logger.info("LEGEND:")
    logger.info("  GREEN  = data from Redshift DB")
    logger.info("  ORANGE = field not available in DB (blank)")
    logger.info("  BLUE   = hardcoded / derived value")
    logger.info("=" * 70)


# ── main ─────────────────────────────────────────────────────────────────────
def main() -> None:
    logger = setup_logging()
    logger.info("=" * 70)
    logger.info("ABFL SANCTION FORMAT EXTRACTION  --  %s", TIMESTAMP)
    logger.info("=" * 70)

    lan_list = get_lan_list(logger)
    sql      = load_sql(lan_list)

    db_cfg = load_db_config()
    logger.info("Connecting to Redshift: %s / %s",
                db_cfg.get("host"), db_cfg.get("database"))
    conn = get_connection(db_cfg)

    try:
        df = run_query(conn, sql, logger)
    except Exception as exc:
        logger.error("Query failed: %s", exc)
        conn.rollback()
        conn.close()
        sys.exit(1)
    finally:
        try:
            conn.close()
        except Exception:
            pass

    logger.info("DB connection closed.")

    if df.empty:
        logger.error("No data returned — check LAN list and SQL.")
        sys.exit(1)

    write_excel(df, logger)
    eval_checks(df, logger)
    logger.info("Output: %s", OUTPUT_PATH)

    try:
        os.startfile(str(OUTPUT_PATH))
    except Exception:
        pass


if __name__ == "__main__":
    main()
