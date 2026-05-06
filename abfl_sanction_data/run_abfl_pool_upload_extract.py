"""
ABFL Pool Upload Format — Full Data Extraction
===============================================
Pulls real data from Redshift for all 297 HE loans and writes
to an Excel file matching the Pool Upload Format sheet structure.

Columns that CANNOT be sourced from the DB are:
  - Left blank in the output
  - Header cell highlighted in ORANGE

Run:
    python ABFL_Formats/run_abfl_pool_upload_extract.py
"""
from __future__ import annotations

import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────────────────
ROOT          = Path(__file__).resolve().parents[1]
ABFL_DIR      = ROOT / "ABFL_Formats"
SOURCE_XLSM   = ABFL_DIR / "Pool Upload Format - Final_.xlsm"
SQL_FILE      = ABFL_DIR / "sql_abfl_pool_upload.sql"
CONFIG_PATH   = ROOT / "config" / "database.yaml"
TIMESTAMP     = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_PATH   = ABFL_DIR / f"HE_pool_upload_FILLED_{TIMESTAMP}.xlsx"
LOG_DIR       = ROOT / "logs"

# ── colours ────────────────────────────────────────────────────────────────
FILL_ORANGE   = PatternFill("solid", fgColor="FF8C00")   # cannot-find cols
FILL_GREEN    = PatternFill("solid", fgColor="C6EFCE")   # filled cols
FILL_BLUE     = PatternFill("solid", fgColor="BDD7EE")   # stub cols (hardcoded)
FONT_WHITE    = Font(color="FFFFFF", bold=True)
FONT_DARK     = Font(color="1F3864", bold=True)

# ── columns that are intentionally blank (no DB source) ────────────────────
# These get ORANGE header highlight.
NOT_FOUND_COLS: dict[str, list[str]] = {
    "Deal and Loan data": [
        "Relationship officer/executive",
        "Interest Compounding Frequency",
        "Credit Period(in days)",
        "Business Type",
        "Anchor Id", "Dealer Id", "Vendor Id",
        "CRM NO", "Card_No", "field_1", "los_scheme_name", "merchant_id",
    ],
    "Customer Data": [
        "Group Type", "Group Name",
        "Residential Status", "Citizenship", "Registration no.",
        "TIN No", "Existing Customer",
        "Customer Risk Category", "Insurance Flag",
        "TAN_NO",
    ],
    "Address Details": [
        "State_ID", "District_ID", "Tehsil",
        "Landmark", "Area",
        "Address Category", "Ownership Type",
        "Residence Duration", "District",
    ],
    "Reference Details": [
        "First Name ", "Middle Name", "Last Name ",
        "Relationship", "Mobile No", "LandLine No",
        "Knowing Since (In Year)", "Address",
    ],
    "Bank Details": ["MICR Code"],
    "Management Details": [
        "Salutation", "StakeHolder Name", "Management Type",
        "Date of Birth(DD-MM-YYYY)", "Email ID", "Mobile No", "PAN No.",
    ],
    "AssetCollateralDetails": [
        "Product Type", "Tax Amount", "Invoice Amount",
        "Invoice No", "Invoice Location", "Due Date DD-MM-YYYY",
        "Machine Description", "Machine Cost", "Discount",
        "Machine Value", "Machine Security Margin",
        "Machine Make", "Machine Model", "Machine Type",
        "Machinery Owner", "Year Of Manufacturing",
        "Identification Number", "Asset Nature",
        "Manufacturer", "Supplier", "Invoice Date DD_MM_YYYY",
        "Security Type", "Address Line2", "Address Line3",
        "Tehsil", "Standard",
        "Property Title", "Property Construction",
        "Document Value", "Technical Val1", "Technical Val2",
        "Valuation MethodId", "Collateral Security Margin",
        "Asset Level", "Additional Construction",
        "Machine Collateral Cost", "Asset Collateral Class",
        "Reff Asset Id", "Asset Insurance",
    ],
    "UploadChargeDetails": [
        "Application Form No.", "Charge Type", "Charge Code",
        "Business Partner Type", "Business Partner Name",
        "Tax Inclusive", "Tax Rate1", "Tax Rate2",
        "Charge Amount", "Final Amount",
        "Charge Calculated On", "Charge Method",
    ],
    "Upload Disbursal Data": [
        "Maker Remarks", "Author Remarks", "Disbursal To",
        "Adjust Total Payable", "Adjust Total Receivable",
        "Loan Curtailment", "Pay To", "Payee Name",
        "Instrument No", "Instrument Date DD-MM-YYYY",
        "Bank Id", "Branch Id", "MICR Code",
        "Tds Amount", "Remarks", "Payment Flag",
    ],
    "Upload Disbursal Schedule": [],
    "Upload Installment Plan": [],
}

# ── SHEET → SQL QUERY KEY mapping ──────────────────────────────────────────
SHEET_QUERY_KEY: dict[str, str | None] = {
    "Deal and Loan data":     "deal",
    "Customer Data":          "customer",
    "Address Details":        "address",
    "Reference Details":      "reference",
    "Bank Details":           "bank",
    "Management Details":     "management",
    "AssetCollateralDetails": "asset",
    "UploadChargeDetails":    "charge",
    "Upload Disbursal Data":  "disbursal",
    "Upload Disbursal Schedule": "disbursal_schedule",
    "Upload Installment Plan": "installment",
    "Master":                 None,   # copied as-is
}

# header row in the source xlsm (0-based) for each sheet
SHEET_HEADER_ROW: dict[str, int | None] = {
    "Deal and Loan data":        2,
    "Customer Data":             1,
    "Address Details":           1,
    "Reference Details":         0,
    "Bank Details":              0,
    "Management Details":        0,
    "AssetCollateralDetails":    1,
    "UploadChargeDetails":       0,
    "Upload Disbursal Data":     1,
    "Upload Disbursal Schedule": 0,
    "Upload Installment Plan":   1,
    "Master":                    None,
}


# ── logging ────────────────────────────────────────────────────────────────
def setup_logging() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f"abfl_pool_upload_{TIMESTAMP}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return logging.getLogger(__name__)


# ── DB helpers ─────────────────────────────────────────────────────────────
def load_db_config() -> dict:
    with CONFIG_PATH.open() as f:
        cfg = yaml.safe_load(f) or {}
    return cfg.get("database") or {}


def get_connection(cfg: dict):
    import redshift_connector
    return redshift_connector.connect(
        host=cfg["host"],
        port=int(cfg["port"]),
        database=cfg["database"],
        user=cfg["user"],
        password=cfg["password"],
        timeout=1800,
    )


def run_query(conn, sql: str, label: str, logger: logging.Logger) -> pd.DataFrame:
    logger.info("Running query: %s ...", label)
    cur = conn.cursor()
    cur.execute(sql)
    cols = [desc[0] for desc in cur.description]
    rows = cur.fetchall()
    cur.close()
    df = pd.DataFrame(rows, columns=cols)
    logger.info("  -> %d rows, %d cols", len(df), len(df.columns))
    return df


# ── SQL parsing ────────────────────────────────────────────────────────────
def parse_queries(sql_text: str, app_list_sql: str) -> dict[str, str]:
    """
    Split sql_abfl_pool_upload.sql into individual queries by
    the  -- [QUERY: <key>]  markers.
    """
    pattern = re.compile(r"--\s*\[QUERY:\s*(\w+)\]", re.IGNORECASE)
    parts   = pattern.split(sql_text)
    queries: dict[str, str] = {}
    # parts = [pre, key1, body1, key2, body2, ...]
    it = iter(parts)
    next(it)  # skip preamble before first marker
    for key in it:
        body = next(it, "")
        # strip trailing separator lines and next-query comments
        body = re.sub(r"--\s*={5,}.*", "", body, flags=re.DOTALL).strip()
        body = body.replace("{LAN_LIST}", app_list_sql).replace("{APP_LIST}", app_list_sql)
        queries[key.strip()] = body
    return queries


# ── Excel helpers ──────────────────────────────────────────────────────────
def col_letter(n: int) -> str:
    return get_column_letter(n)


def highlight_headers(ws, df_cols: list[str], not_found: list[str]) -> None:
    """
    Row 1 of the sheet = header row written by to_excel.
    Green  = data present / stub hardcoded.
    Orange = field not available in DB.
    """
    nf_lower = {c.strip().lower() for c in not_found}
    for idx, col in enumerate(df_cols, start=1):
        cell = ws.cell(row=1, column=idx)
        if col.strip().lower() in nf_lower:
            cell.fill = FILL_ORANGE
            cell.font = FONT_WHITE
        else:
            cell.fill = FILL_GREEN
            cell.font = FONT_DARK


# ── app list ───────────────────────────────────────────────────────────────
def get_app_list(logger: logging.Logger) -> tuple[list[str], str]:
    logger.info("Reading application list from source xlsm …")
    deal = pd.read_excel(SOURCE_XLSM, sheet_name="Deal and Loan data",
                         header=2, engine="openpyxl")
    def norm(x):
        if pd.isna(x): return None
        if isinstance(x, float): return str(int(x))
        return str(x).strip()
    apps = [norm(v) for v in deal["Application Form No."].dropna().unique()]
    apps = [a for a in apps if a]
    sql_list = ", ".join(f"'{a}'" for a in apps)
    logger.info("  %d HE application numbers", len(apps))
    return apps, sql_list


# ── main ───────────────────────────────────────────────────────────────────
def main() -> None:
    logger = setup_logging()
    logger.info("=" * 70)
    logger.info("ABFL POOL UPLOAD EXTRACTION  --  %s", TIMESTAMP)
    logger.info("=" * 70)

    apps, app_list_sql = get_app_list(logger)  # apps are sz_loan_account_no (LANs)

    sql_text = SQL_FILE.read_text(encoding="utf-8")
    queries  = parse_queries(sql_text, app_list_sql)
    logger.info("Parsed %d SQL queries: %s", len(queries), list(queries.keys()))

    db_cfg = load_db_config()
    logger.info("Connecting to Redshift: %s / %s", db_cfg.get("host"), db_cfg.get("database"))
    conn = get_connection(db_cfg)

    results: dict[str, pd.DataFrame] = {}
    try:
        for key, sql in queries.items():
            if not sql.strip():
                logger.warning("Empty SQL for key '%s' — skipping", key)
                continue
            try:
                results[key] = run_query(conn, sql, key, logger)
            except Exception as exc:
                logger.error("Query '%s' failed: %s", key, exc)
                results[key] = pd.DataFrame()
                # rollback aborted transaction so next query can run
                try:
                    conn.rollback()
                except Exception:
                    pass
    finally:
        conn.close()
        logger.info("DB connection closed.")

    # read Master sheet as-is from source
    master_raw = pd.read_excel(SOURCE_XLSM, sheet_name="Master",
                               header=None, engine="openpyxl")

    logger.info("Writing output Excel -> %s", OUTPUT_PATH)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:

        for sheet_name, query_key in SHEET_QUERY_KEY.items():
            if query_key is None:
                # Master — copy raw
                master_raw.to_excel(writer, sheet_name="Master",
                                    header=False, index=False)
                continue

            hdr_row = SHEET_HEADER_ROW.get(sheet_name, 0)
            template_df = pd.read_excel(
                SOURCE_XLSM, sheet_name=sheet_name,
                header=hdr_row, engine="openpyxl", nrows=0
            )
            template_cols = template_df.columns.tolist()

            df_db = results.get(query_key, pd.DataFrame())

            if df_db.empty:
                # write header only
                out_df = pd.DataFrame(columns=template_cols)
            else:
                # align to template columns
                out_df = pd.DataFrame(columns=template_cols)
                for tc in template_cols:
                    tc_lower = tc.strip().lower()
                    # exact match first, then case-insensitive
                    matched = None
                    for dc in df_db.columns:
                        if dc == tc:
                            matched = dc; break
                    if matched is None:
                        for dc in df_db.columns:
                            if dc.strip().lower() == tc_lower:
                                matched = dc; break
                    if matched:
                        out_df[tc] = df_db[matched].values[:len(df_db)]

            out_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

            # apply header highlights
            ws = writer.sheets[sheet_name[:31]]
            nf = NOT_FOUND_COLS.get(sheet_name, [])
            highlight_headers(ws, template_cols, nf)

    # ── eval checks ────────────────────────────────────────────────────────
    logger.info("=" * 70)
    logger.info("EVAL CHECKS")
    logger.info("=" * 70)
    for key, df in results.items():
        logger.info("%-26s rows: %d", key, len(df))
        if not df.empty:
            # null check on first column (usually Application Form No.)
            c0 = df.columns[0]
            nulls = df[c0].isna().sum()
            if nulls:
                logger.warning("  %s — %d null values in '%s'", key, nulls, c0)
            # duplicate check
            dups = df.duplicated(subset=[c0]).sum() if len(df.columns) >= 1 else 0
            logger.info("  Duplicate '%s': %d", c0, dups)

    logger.info("=" * 70)
    logger.info("LEGEND (header colours in output Excel):")
    logger.info("  GREEN  = data pulled from Redshift DB")
    logger.info("  ORANGE = field NOT available in DB (left blank)")
    logger.info("  BLUE   = hardcoded/stub value")
    logger.info("Output: %s", OUTPUT_PATH)

    try:
        import os
        os.startfile(str(OUTPUT_PATH))
    except Exception:
        pass


if __name__ == "__main__":
    main()
