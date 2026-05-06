"""
Strong Column-Level Eval — ABFL Sanction Format
Works on already-extracted Excel output (HE 297 + DA Pool bulk).
50+ checks: fill rate, enum, range, format, cross-column logic, DPD/Bounce patterns.
"""
import re
import sys
import glob
import pandas as pd
from pathlib import Path
from datetime import datetime, date

ABFL_DIR = Path(__file__).parent
TODAY    = pd.Timestamp(date.today())

# ─── Load latest files ────────────────────────────────────────────────────────
he_files   = sorted(glob.glob(str(ABFL_DIR / "HE_sanction_format_*.xlsx")))
bulk_files = sorted(glob.glob(str(ABFL_DIR / "DA_Pool_sanction_format_*.xlsx")))

if not he_files and not bulk_files:
    sys.exit("No output files found.")

targets = []
if he_files:   targets.append(("HE-297",  he_files[-1]))
if bulk_files: targets.append(("DA-POOL", bulk_files[-1]))

# ─── Helpers ──────────────────────────────────────────────────────────────────
def gcol(df, substr):
    s = substr.lower()
    return next((c for c in df.columns if s in c.lower()), None)

def filled(series):
    return series.notna() & (series.astype(str).str.strip().isin(['', 'nan', 'None', 'NaT']) == False)

results = []

def chk(tag, col_no, col_name, check_name, passed, detail="", warn=False):
    status = "PASS" if passed else ("WARN" if warn else "FAIL")
    results.append({
        "Tag": tag, "Col#": col_no, "Column": str(col_name)[:52],
        "Check": check_name, "Status": status, "Detail": detail
    })

# ─── Run for each file ────────────────────────────────────────────────────────
NOT_IN_DB = {
    "Subvention Scheme if any", "NRI Loan (Yes/No)", "ECLGS",
    "Link loan Part of pool (Yes/ No)", "Link Loan/Top up loan Number if applicable",
    "Legal Report", "PMAY subsidy status"
}

# Columns where partial fill is structurally correct (blank = valid for certain rows)
# NOTE: "in case of self employed" is NOT here — account format requires Salaried/SENP/SEP
#       to be filled for ALL borrowers, not just SE. No conditional nullification.
PARTIAL_OK = {
    "pdd status, if pending",     # blank when PDD complete
    "cersai registration date",   # not always registered
}

for tag, fpath in targets:
    print(f"\nLoading {tag}: {Path(fpath).name}")
    df = pd.read_excel(fpath, dtype=str)
    N  = len(df)
    print(f"  {N} rows x {len(df.columns)} columns")

    # ── A. FILL RATE ALL COLUMNS ─────────────────────────────────────────────
    for i, c in enumerate(df.columns, 1):
        col_data = filled(df.iloc[:, i - 1])
        pct = round(col_data.sum() / N * 100, 1)
        # NOT_IN_DB: substring match to handle columns with \n or long suffixes
        if any(key.lower() in c.lower() for key in NOT_IN_DB):
            chk(tag, i, c, "Fill Rate", True, f"No DB source (expected blank) {col_data.sum()}/{N}")
        elif any(p in c.lower() for p in PARTIAL_OK):
            chk(tag, i, c, "Fill Rate", True,
                f"{col_data.sum()}/{N} ({pct}%) — partial fill expected by design")
        else:
            passed = col_data.sum() > 0
            # Secondary applicants: applicants 3-9 + pandas-deduplicated PAN (.2-.7)
            is_secondary = any(x in c.lower() for x in ["applicant 3","applicant 4","applicant 5",
                                                          "applicant 6","applicant 7","applicant 8",
                                                          "applicant 9"]) or \
                           any(c.lower().startswith(f"pan number.{n}") for n in range(2, 8))
            chk(tag, i, c, "Fill Rate", passed if not is_secondary else col_data.sum() > 0,
                f"{col_data.sum()}/{N} ({pct}%)",
                warn=(passed and pct < 50 and is_secondary))

    # ── B. DUPLICATE Loan No. ────────────────────────────────────────────────
    c = gcol(df, "loan no")
    if c:
        dups = df[c].duplicated().sum()
        chk(tag, 2, c, "No Duplicates", dups == 0, f"{dups} duplicate loan nos")

    # ── C. CUSTOMER SUB-TYPE (must be non-null for applicant 1 & 2) ──────────
    for seq, col_no in [("primary", 10), ("applicant 2)", 21)]:
        c = gcol(df, f"customer sub-type ({seq}" if seq == "primary" else "sub-type (applicant 2")
        if not c:
            c = gcol(df, seq.split(")")[0] if ")" in seq else seq)
        if c:
            n_filled = filled(df[c]).sum()
            chk(tag, col_no, c, "Customer Sub-type Non-null", n_filled == N,
                f"{N - n_filled} missing")

    # ── D. CONSTITUTION ENUM ─────────────────────────────────────────────────
    c = gcol(df, "constitution")
    if c:
        valid_const = {"Salaried", "Sole Proprietor", "Partnership", "Private Limited",
                       "Public Limited", "LLP", "HUF", "Trust", "NGO", "Society",
                       "Co-operative Society", "Individual", "Non-Individual"}
        vals = df[c].dropna().astype(str).str.strip()
        bad  = vals[~vals.isin(valid_const)]
        chk(tag, 13, c, "Constitution Enum", len(bad) == 0,
            f"OK values: {vals.value_counts().head(4).to_dict()}" +
            (f" | Unknown: {bad.value_counts().head(5).to_dict()}" if len(bad) else ""))

    # ── C2. SENP/SEP col — account format: ALL borrowers must be filled (Salaried/SENP/SEP) ──
    c_senp = gcol(df, "in case of self employed")
    if c_senp:
        n_filled = filled(df[c_senp]).sum()
        vals = df[c_senp].dropna().astype(str).str.strip().value_counts().to_dict()
        chk(tag, 11, c_senp, "SENP/SEP 100% filled (account format)",
            n_filled == N,
            f"{n_filled}/{N} | dist={vals}")

    # ── E. INCOME ASSESSMENT METHOD non-null ────────────────────────────────
    c = gcol(df, "income assesment")
    if c:
        n = filled(df[c]).sum()
        chk(tag, 70, c, "Income Method >= 95% filled", n / N >= 0.95, f"{n}/{N} ({n/N*100:.1f}%)")

    # ── F. LTV AT SANCTION (1-100%) ──────────────────────────────────────────
    c = gcol(df, "time of original sanction")
    if c:
        ltv = pd.to_numeric(df[c], errors="coerce")
        n_filled = ltv.notna().sum()
        bad_range = ltv[(ltv < 0) | (ltv > 105)].dropna()
        chk(tag, 92, c, "LTV Fill >= 90%", n_filled / N >= 0.90,
            f"{n_filled}/{N} ({n_filled/N*100:.1f}%) filled")
        chk(tag, 92, c, "LTV Range (0-105%)", len(bad_range) == 0,
            f"min={ltv.min():.1f} max={ltv.max():.1f} out-of-range={len(bad_range)}")

    # ── G. CURRENT LTV (0-150%) ──────────────────────────────────────────────
    c = gcol(df, "current ltv")
    if c:
        ltv = pd.to_numeric(df[c], errors="coerce").dropna()
        bad = ltv[(ltv < 0) | (ltv > 200)]
        chk(tag, 91, c, "Current LTV Range (0-200%)", len(bad) == 0,
            f"min={ltv.min():.1f} max={ltv.max():.1f} bad={len(bad)}")

    # ── H. FOIR in % (0-100) ────────────────────────────────────────────────
    c = gcol(df, "foir")
    if c:
        foir = pd.to_numeric(df[c], errors="coerce").dropna()
        bad  = foir[(foir < 0) | (foir > 100)]
        chk(tag, 128, c, "FOIR Range (0-100%)", len(bad) == 0,
            f"min={foir.min():.1f} max={foir.max():.1f} bad={len(bad)}")
        chk(tag, 128, c, "FOIR Fill >= 90%", len(foir) / N >= 0.90,
            f"{len(foir)}/{N} ({len(foir)/N*100:.1f}%) filled")

    # ── I. CURRENT ROI (1-40%) ───────────────────────────────────────────────
    c = gcol(df, "current roi")
    if c:
        roi = pd.to_numeric(df[c], errors="coerce").dropna()
        bad = roi[(roi < 0.5) | (roi > 40)]
        chk(tag, 81, c, "ROI Range (0.5-40%)", len(bad) == 0,
            f"min={roi.min():.2f} max={roi.max():.2f} bad={len(bad)}")

    # ── J. RATE TYPE ENUM ────────────────────────────────────────────────────
    c = gcol(df, "rate type")
    if c:
        vals = df[c].dropna().astype(str).str.strip().str.upper()
        valid = {"FIXED", "F", "VARIABLE", "V", "FLOATING", "SEMI-FIXED", "M",
                 "SEMI FIXED", "MIXED", "ADJUSTABLE"}
        bad   = vals[~vals.isin(valid)]
        chk(tag, 80, c, "Rate Type Enum", len(bad) == 0,
            f"values: {vals.value_counts().to_dict()}" +
            (f" | unknown: {bad.value_counts().to_dict()}" if len(bad) else ""))

    # ── K. SANCTION AMOUNT (3L – 2Cr) ───────────────────────────────────────
    c = gcol(df, "sanctioned amount")
    if c:
        amt = pd.to_numeric(df[c], errors="coerce")
        low  = (amt < 300000).sum()
        high = (amt > 20000000).sum()
        chk(tag, 76, c, "Sanction Amt 3L-2Cr", low == 0 and high == 0,
            f"min={amt.min():,.0f} max={amt.max():,.0f} <3L:{low} >2Cr:{high}")

    # ── L. PRINCIPAL O/S >= 0 ────────────────────────────────────────────────
    c = gcol(df, "principal o/s")
    if c:
        pos = pd.to_numeric(df[c], errors="coerce")
        neg = (pos < 0).sum()
        chk(tag, 82, c, "Principal O/S >= 0", neg == 0,
            f"min={pos.min():,.0f} max={pos.max():,.0f} neg={neg}")

    # ── M. DISBURSED <= SANCTIONED ───────────────────────────────────────────
    c_d = gcol(df, "disbursed amount")
    c_s = gcol(df, "sanctioned amount")
    if c_d and c_s:
        disb = pd.to_numeric(df[c_d], errors="coerce")
        sanc = pd.to_numeric(df[c_s], errors="coerce")
        over = (disb > sanc * 1.05).sum()
        chk(tag, 77, "Disbursed vs Sanctioned", "Disbursed <= Sanctioned+5%", over == 0,
            f"{over} loans exceeded", warn=(over > 0))

    # ── N. BALANCE TENURE <= SANCTIONED TENURE ───────────────────────────────
    c_b = gcol(df, "balance tenure")
    c_t = gcol(df, "sanctioned tenure")
    if c_b and c_t:
        bal = pd.to_numeric(df[c_b], errors="coerce")
        ten = pd.to_numeric(df[c_t], errors="coerce")
        bad = (bal > ten + 1).sum()
        chk(tag, 79, "Tenure Cross-check", "Balance Tenure <= Sanctioned Tenure", bad == 0,
            f"{bad} violations")

    # ── O. AGE (18-80) ───────────────────────────────────────────────────────
    c = gcol(df, "financial applicant age")
    if c:
        age = pd.to_numeric(df[c], errors="coerce").dropna()
        bad = age[(age < 18) | (age > 80)]
        chk(tag, 127, c, "Age Range (18-80 yrs)", len(bad) == 0,
            f"min={age.min():.0f} max={age.max():.0f} bad={len(bad)}")

    # ── P. DOB / DATE SANITY ─────────────────────────────────────────────────
    c = gcol(df, "dob/doi\n(primary")
    if not c: c = gcol(df, "dob/doi")
    if c:
        dob = pd.to_datetime(df[c], errors="coerce")
        future = (dob > TODAY).sum()
        old    = (dob < pd.Timestamp("1930-01-01")).sum()
        chk(tag, 14, c, "DOB Sanity (1930-today)", future == 0 and old == 0,
            f"future={future} pre1930={old} null={dob.isna().sum()}")

    # ── Q. PAN FORMAT (XXXXX9999X) ───────────────────────────────────────────
    pan_cols = [(i+1, c) for i, c in enumerate(df.columns) if "pan" in c.lower()]
    for col_no, c in pan_cols:
        # Skip deduplicated secondary-applicant PAN columns (.2 through .7)
        if any(c.lower().startswith(f"pan number.{n}") for n in range(2, 8)):
            continue
        pans = df[c].dropna().astype(str).str.strip()
        pans = pans[pans != 'nan']
        if len(pans) == 0: continue
        bad = pans[~pans.str.match(r"^[A-Z]{5}[0-9]{4}[A-Z]$")]
        chk(tag, col_no, c, "PAN Format", len(bad) == 0,
            f"{len(pans)} filled, {len(bad)} invalid" +
            (f": {bad.head(3).tolist()}" if len(bad) else ""))

    # ── R. CIBIL SCORE (300-900 or -1) ──────────────────────────────────────
    for substr, col_no in [("originated cibil", 106), ("current cibil", 107)]:
        c = gcol(df, substr)
        if c:
            scores = pd.to_numeric(df[c], errors="coerce").dropna()
            bad    = scores[~((scores >= 300) & (scores <= 900)) & (scores != -1)]
            chk(tag, col_no, c, "CIBIL (300-900 or -1)", len(bad) == 0,
                f"n={len(scores)} null={df[c].isna().sum()} bad={len(bad)}")

    # ── S. DPD STRING SINCE INCEPTION (format: 000-000-... or STATUS words) ──
    c = gcol(df, "dpd string (since inception)")
    if c:
        vals = df[c].dropna().astype(str)
        bad = vals[~vals.str.match(r"^(\d{3}(-\d{3})*|[A-Z][A-Z0-9_\- ]*)$")]
        chk(tag, 122, c, "DPD Inception Format (NNN-NNN or STATUS)", len(bad) == 0,
            f"{len(vals)} filled, {len(bad)} malformed" +
            (f": {bad.head(3).tolist()}" if len(bad) else ""))

    # ── T. DPD STRING LAST 12M ───────────────────────────────────────────────
    c = gcol(df, "dpd string - last 12")
    if c:
        vals = df[c].dropna().astype(str)
        bad  = vals[~vals.str.match(r"^\d{3}(-\d{3})*$")]
        chk(tag, 121, c, "DPD 12M Format (NNN-NNN)", len(bad) == 0,
            f"{len(vals)} filled, {len(bad)} malformed")

    # ── U. MAX EVER DPD >= 0 ────────────────────────────────────────────────
    c = gcol(df, "max ever dpd")
    if c:
        dpd = pd.to_numeric(df[c], errors="coerce").dropna()
        neg = (dpd < 0).sum()
        chk(tag, 126, c, "Max Ever DPD >= 0", neg == 0,
            f"min={dpd.min():.0f} max={dpd.max():.0f} neg={neg}")

    # ── V. BOUNCE STRINGS (B/C pattern) ────────────────────────────────────
    for substr, col_no in [
        ("bounce dpd string for last 12", 123),
        ("gross bounce string", 124),
        ("bounce string since last 12", 125),
    ]:
        c = gcol(df, substr)
        if c:
            vals = df[c].dropna().astype(str)
            bad  = vals[~vals.str.match(r"^[BC](-[BC])*$")]
            chk(tag, col_no, c, "Bounce Pattern (B/C)", len(bad) == 0,
                f"{len(vals)} filled, {len(bad)} malformed" +
                (f": {bad.head(3).tolist()}" if len(bad) else ""))

    # ── W. REPAYMENT MODE ENUM ───────────────────────────────────────────────
    c = gcol(df, "repayment mode")
    if c:
        vals = df[c].dropna().astype(str).str.strip().str.upper()
        valid = {"NACH", "NEFT", "PDC", "SI", "ACH", "ECS", "CASH",
                 "STANDING INSTRUCTION", "DIRECT DEBIT", "CHEQUE",
                 "ESCROW", "UPI", "ONLINE"}
        bad   = vals[~vals.isin(valid)]
        chk(tag, None, c, "Repayment Mode Enum", len(bad) == 0,
            f"values: {vals.value_counts().to_dict()}" +
            (f" | unknown: {bad.value_counts().to_dict()}" if len(bad) else ""))

    # ── X. PDD STATUS (Y/N) ──────────────────────────────────────────────────
    c = gcol(df, "pdd status complete")
    if c:
        vals = df[c].dropna().astype(str).str.strip().str.upper()
        bad  = vals[~vals.isin({"Y", "N"})]
        chk(tag, 108, c, "PDD Status (Y/N)", len(bad) == 0,
            f"Y={( vals == 'Y').sum()} N={(vals == 'N').sum()} bad={len(bad)}")

    # ── Y. NPA FLAG (Y/N) ────────────────────────────────────────────────────
    c = gcol(df, "loan became npa")
    if c:
        vals = df[c].dropna().astype(str).str.strip().str.upper()
        bad  = vals[~vals.isin({"Y", "N"})]
        chk(tag, 119, c, "NPA Flag (Y/N)", len(bad) == 0,
            f"Y={(vals=='Y').sum()} N={(vals=='N').sum()} bad={len(bad)}")

    # ── Z. RESTRUCTURED (Y/N) ────────────────────────────────────────────────
    c = gcol(df, "restructured flag")
    if c:
        vals = df[c].dropna().astype(str).str.strip().str.upper()
        bad  = vals[~vals.isin({"Y", "N"})]
        chk(tag, 115, c, "Restructured (Y/N)", len(bad) == 0,
            f"values: {vals.value_counts().to_dict()}")

    # ── AA. COLLATERAL VALUE > 0 ─────────────────────────────────────────────
    c = gcol(df, "collateral  value")
    if not c: c = gcol(df, "collateral value")
    if c:
        cv = pd.to_numeric(df[c], errors="coerce")
        bad = (cv <= 0).sum()
        chk(tag, 98, c, "Collateral Value > 0", bad == 0,
            f"min={cv.min():,.0f} max={cv.max():,.0f} zero/neg={bad}")

    # ── AB. COLLATERAL DESC non-null ────────────────────────────────────────
    c = gcol(df, "collateral description")
    if c:
        n = filled(df[c]).sum()
        chk(tag, None, c, "Collateral Desc >= 90% filled", n / N >= 0.90,
            f"{n}/{N} ({n/N*100:.1f}%)")

    # ── AC. END USE fill rate >= 80% ────────────────────────────────────────
    c = gcol(df, "end use")
    if c:
        n = filled(df[c]).sum()
        chk(tag, 67, c, "End Use >= 80% filled", n / N >= 0.80,
            f"{n}/{N} ({n/N*100:.1f}%)")

    # ── AD. PINCODE 6-digit ──────────────────────────────────────────────────
    for substr, col_no in [("zip code", 18), ("pincode of property", 103)]:
        c = gcol(df, substr)
        if c:
            pins = df[c].dropna().astype(str).str.strip()
            bad  = pins[~pins.str.match(r"^\d{6}$")]
            chk(tag, col_no, c, "Pincode (6 digits)", len(bad) == 0,
                f"{len(bad)} invalid: {bad.head(3).tolist()}")

    # ── AE. SANCTION DATE (2010-today) ───────────────────────────────────────
    c = gcol(df, "sanctioned date")
    if c:
        dates  = pd.to_datetime(df[c], errors="coerce")
        future = (dates > TODAY).sum()
        old    = (dates < pd.Timestamp("2010-01-01")).sum()
        chk(tag, 71, c, "Sanction Date (2010-today)", future == 0 and old == 0,
            f"future={future} pre-2010={old} null={dates.isna().sum()}")

    # ── AF. FIRST DISB <= LAST DISB ─────────────────────────────────────────
    c_first = gcol(df, "first disbursement")
    c_last  = gcol(df, "last disbursement")
    if c_first and c_last:
        d1 = pd.to_datetime(df[c_first], errors="coerce")
        d2 = pd.to_datetime(df[c_last],  errors="coerce")
        bad = (d1 > d2).sum()
        chk(tag, 72, "Disb Dates", "First Disb <= Last Disb", bad == 0,
            f"{bad} violations")

    # ── AG. PSL/NPSL ENUM ───────────────────────────────────────────────────
    c = gcol(df, "psl/npsl")
    if c:
        vals = df[c].dropna().astype(str).str.strip().str.upper()
        valid = {"P", "NP", "N", "PSL", "NPSL", "Y"}
        bad   = vals[~vals.isin(valid)]
        chk(tag, 120, c, "PSL/NPSL Enum", len(bad) == 0,
            f"values: {vals.value_counts().to_dict()}")

    # ── AH. ANNUAL INCOME > 0 ────────────────────────────────────────────────
    c = gcol(df, "annual income")
    if c:
        inc = pd.to_numeric(df[c], errors="coerce")
        bad = (inc <= 0).sum()
        chk(tag, 129, c, "Annual Income > 0", bad == 0,
            f"min={inc.min():,.0f} max={inc.max():,.0f} zero/neg={bad}")

    # ── AI. GENDER ENUM ──────────────────────────────────────────────────────
    c = gcol(df, "gender for primary")
    if c:
        vals = df[c].dropna().astype(str).str.strip().str.upper()
        valid = {"M", "F", "O", "MALE", "FEMALE", "OTHER"}
        bad   = vals[~vals.isin(valid)]
        chk(tag, 7, c, "Gender Enum (M/F/O/MALE/FEMALE)", len(bad) == 0,
            f"values: {vals.value_counts().to_dict()}")

    # ── AJ. UNDER CONSTRUCTION (Y/N) ────────────────────────────────────────
    c = gcol(df, "under construction")
    if c:
        vals = df[c].dropna().astype(str).str.strip().str.upper()
        bad  = vals[~vals.isin({"Y", "N", "YES", "NO"})]
        chk(tag, 96, c, "Under Construction (Y/N)", len(bad) == 0,
            f"values: {vals.value_counts().to_dict()}")

    # ── AK. CROSS CHECK: Overdue principal aligns with DPD ──────────────────
    c_od  = gcol(df, "overdue amt")
    c_dpd = gcol(df, "current dpd")
    if c_od and c_dpd:
        od  = pd.to_numeric(df[c_od],  errors="coerce").fillna(0)
        dpd = pd.to_numeric(df[c_dpd], errors="coerce").fillna(0)
        mismatch = ((od > 0) & (dpd == 0)).sum()
        chk(tag, "X", "Overdue vs DPD", "Overdue>0 implies DPD>0", mismatch == 0,
            f"{mismatch} cases with overdue>0 but DPD=0", warn=(mismatch > 0))

    # ── AL. CERSAI ID fill rate ──────────────────────────────────────────────
    c = gcol(df, "cersai")
    if c:
        n = filled(df[c]).sum()
        chk(tag, 131, c, "CERSAI ID >= 85% filled", n / N >= 0.85,
            f"{n}/{N} ({n/N*100:.1f}%)")

    # ── AM. SUMMARY ROW COUNT ────────────────────────────────────────────────
    chk(tag, 0, "DATASET", "Row Count > 0", N > 0, f"{N} rows")


# ─── Render report ────────────────────────────────────────────────────────────
rdf = pd.DataFrame(results)
rdf["Status"] = pd.Categorical(rdf["Status"], ["FAIL","WARN","PASS"], ordered=True)

ts = datetime.now().strftime("%Y-%m-%d %H:%M")
print("\n" + "=" * 100)
print(f"  STRONG EVAL REPORT  —  {ts}")
print("=" * 100)

for tag in rdf["Tag"].unique():
    sub  = rdf[rdf["Tag"] == tag]
    fails = (sub["Status"] == "FAIL").sum()
    warns = (sub["Status"] == "WARN").sum()
    passes= (sub["Status"] == "PASS").sum()
    score = round(passes / len(sub) * 100, 1)

    print(f"\n{'-'*100}")
    print(f"  [{tag}]   Total: {len(sub)}  |  PASS: {passes}  |  WARN: {warns}  |  FAIL: {fails}  |  Score: {score}%")
    print(f"{'-'*100}")

    for status, marker in [("FAIL","FAIL"), ("WARN","WARN"), ("PASS","PASS")]:
        grp = sub[sub["Status"] == status].sort_values("Col#")
        if grp.empty: continue
        print(f"\n  [{marker}] ({len(grp)} checks)")
        for _, row in grp.iterrows():
            col_lbl = f"[C{str(row['Col#']):>3}]" if str(row["Col#"]) not in ("None", "0", "X", "") else "[   ]"
            print(f"    {col_lbl}  {str(row['Column']):<52}  {str(row['Check']):<45}  {str(row['Detail'])[:80]}")

# ─── Save to Excel ────────────────────────────────────────────────────────────
ts_fn = datetime.now().strftime("%Y%m%d_%H%M%S")
out   = ABFL_DIR / f"_eval_strong_{ts_fn}.xlsx"

with pd.ExcelWriter(out, engine="openpyxl") as writer:
    for tag in rdf["Tag"].unique():
        sub = rdf[rdf["Tag"] == tag].copy()
        sub.to_excel(writer, sheet_name=tag, index=False)
        ws = writer.sheets[tag]
        from openpyxl.styles import PatternFill, Font
        fill_fail = PatternFill("solid", fgColor="FF4444")
        fill_warn = PatternFill("solid", fgColor="FFA500")
        fill_pass = PatternFill("solid", fgColor="00B050")
        font_w    = Font(bold=True, color="FFFFFF")
        col_idx   = list(sub.columns).index("Status") + 1
        for row_idx, status in enumerate(sub["Status"], 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if status == "FAIL":
                cell.fill = fill_fail; cell.font = font_w
            elif status == "WARN":
                cell.fill = fill_warn; cell.font = font_w
            elif status == "PASS":
                cell.fill = fill_pass; cell.font = font_w
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                (len(str(c.value)) for c in col if c.value), default=10) + 2

print(f"\n\nEval saved: {out}")
import os
try: os.startfile(str(out))
except: pass
