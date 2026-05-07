"""
Custom LAN List Sanction Format Extract
Reads LANs from custom_lan_list.txt, extracts 132-column sanction format,
and writes Excel with identical formatting to HE/DA-Pool scripts.
"""
from __future__ import annotations
import logging, os
from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pyxlsb
import redshift_connector

ROOT        = Path(__file__).resolve().parents[1]
ABFL_DIR    = ROOT / "ABFL_Formats"
SQL_FILE    = ABFL_DIR / "sql_abfl_sanction_format.sql"
LAN_FILE    = ABFL_DIR / "custom_lan_list.txt"
TEMPLATE    = ABFL_DIR / "Sancttion_Format.xlsb"
CONFIG_PATH = ROOT / "config" / "database.yaml"
TIMESTAMP   = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_PATH = ABFL_DIR / f"Custom_LAN_sanction_format_{TIMESTAMP}.xlsx"
SHEET_NAME  = "Required for pool shortlisting"

FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
FILL_ORANGE = PatternFill("solid", fgColor="FF8C00")
FILL_BLUE   = PatternFill("solid", fgColor="BDD7EE")
FONT_DARK   = Font(color="1F3864", bold=True)
FONT_WHITE  = Font(color="FFFFFF", bold=True)

NOT_FOUND = {
    "Subvention Scheme if any",
    "PMAY subsidy status \n(Claimed & received/ Claimed & not received / Not claimed/Claimed but rejected)",
    "NRI Loan (Yes/No)",
    "ECLGS",
    "Link loan Part of pool (Yes/ No)",
    "Link Loan/Top up loan Number if applicable",
    "Legal Report ",
}

HARDCODED = {
    "PMAY Flag (Y/N) - All are HE cases thus marked as No",
    "EMI Frequency",
    "Customer Type (Primary Applicant)",
    "Customer Type (Applicant 2)",
    "Customer Type (Applicant 3)",
    "Customer Type (Applicant 4)",
    "Customer Type (Applicant 5)",
    "Customer Type (Applicant 6)",
    "Customer Type (Applicant 7)",
    "Customer Type (Applicant 8)",
    "Customer Type (Applicant 9)",
    "Open Plot (Y/N)",
    "Current LTV at asset level",
    "fully disbursed Yes/No",
    "Staff Loan (Yes/No)",
    "Restructured Flag (including OTR 1/OTR 2)",
    "Loan became NPA since origination (Y/N)",
}

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-6s %(message)s")
log = logging.getLogger(__name__)


def get_template_columns() -> list[str]:
    try:
        with pyxlsb.open_workbook(str(TEMPLATE)) as wb:
            with wb.get_sheet(1) as sheet:
                rows = list(sheet.rows())
        best_row, best_cnt = 0, 0
        for i, row in enumerate(rows[:10]):
            cnt = sum(1 for c in row if c.v is not None and str(c.v).strip())
            if cnt > best_cnt:
                best_cnt, best_row = cnt, i
        cols = [str(c.v).strip() if c.v is not None else "" for c in rows[best_row]]
        cols = [c for c in cols if c]
        log.info("Template columns read: %d from %s", len(cols), TEMPLATE.name)
        return cols
    except Exception as e:
        log.warning("Could not read template: %s — using SQL names", e)
        return []


def load_lans() -> list[str]:
    lans = [l.strip() for l in LAN_FILE.read_text().splitlines() if l.strip()]
    log.info("Loaded %d LANs from %s", len(lans), LAN_FILE.name)
    return lans


def load_sql(lans: list[str]) -> str:
    raw = SQL_FILE.read_text(encoding="utf-8")
    quoted = ", ".join(f"'{l}'" for l in lans)
    parts = raw.split("\n\n\n-- ===")
    return parts[0].strip().replace("{LAN_LIST}", quoted)


def run_query(conn, sql: str) -> pd.DataFrame:
    log.info("Executing SQL (%d chars) ...", len(sql))
    cur = conn.cursor()
    cur.execute(sql)
    cols = [d[0] for d in cur.description]
    rows = cur.fetchall()
    cur.close()
    df = pd.DataFrame(rows, columns=cols)
    log.info("Fetched %d rows x %d columns", len(df), len(df.columns))
    return df


def write_excel(df: pd.DataFrame, template_cols: list[str]) -> None:
    log.info("Writing Excel: %s", OUTPUT_PATH)
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Build headers: use df.columns as-is for any extra cols inserted mid-SQL,
    # and template_cols for the standard 132-col block.
    # The SQL may return extra cols (e.g. "Application No.") inserted right after
    # "Loan No." at position 3. Match by using df.columns directly — they already
    # carry the AS aliases from SQL.
    df_cols = list(df.columns)
    n_cols  = len(df_cols)
    if n_cols <= len(template_cols):
        headers = template_cols[:n_cols]
    else:
        # Extra col(s) are in df_cols but not in template_cols.
        # Find the insertion point by scanning for any df col not in template_cols.
        headers = []
        tmpl_idx = 0
        for dc in df_cols:
            if tmpl_idx < len(template_cols) and dc == template_cols[tmpl_idx]:
                headers.append(template_cols[tmpl_idx])
                tmpl_idx += 1
            else:
                # Extra column (e.g. "Application No.") — keep as-is
                headers.append(dc)

    # Header row
    for ci, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border
        if col in NOT_FOUND:
            cell.fill = FILL_ORANGE
            cell.font = FONT_WHITE
        elif col in HARDCODED:
            cell.fill = FILL_BLUE
            cell.font = FONT_DARK
        else:
            cell.fill = FILL_GREEN
            cell.font = FONT_DARK

    # Data rows
    log.info("Writing %d data rows ...", len(df))
    for ri, row_data in enumerate(df.values, 2):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci,
                           value=None if pd.isna(val) else val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Adaptive column widths
    for ci, col in enumerate(headers, 1):
        col_letter = get_column_letter(ci)
        header_len = max(len(line) for line in str(col).splitlines())
        sample = df.iloc[:50, ci - 1].astype(str).str.len()
        data_len = int(sample.max()) if len(sample) > 0 and pd.notna(sample.max()) else 0
        ws.column_dimensions[col_letter].width = min(max(header_len, data_len, 8), 40)

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_PATH)
    log.info("Saved: %s", OUTPUT_PATH)


def quick_eval(df: pd.DataFrame, template_cols: list[str]) -> None:
    N = len(df)
    headers = template_cols if template_cols else list(df.columns)
    log.info("=" * 65)
    log.info("COMPLETION & SEMANTICS EVAL  (%d loans)", N)
    log.info("=" * 65)

    issues = []
    for i in range(min(len(headers), len(df.columns))):
        col   = headers[i]
        series = df.iloc[:, i]
        filled = series.notna().sum()
        pct    = filled / N * 100 if N else 0
        blank  = N - filled
        enc    = series.dropna().astype(str).str.match(
            r'^[A-Za-z0-9+/]{20,}={0,2}$').sum()
        status = "ENCRYPTED" if enc > 0 else ("LOW" if pct < 30 else "OK")
        if status != "OK" or blank > 0:
            issues.append(dict(col=col[:55], pct=round(pct, 1),
                               blank=blank, enc=enc, status=status))

    issues.sort(key=lambda x: (x["status"] != "ENCRYPTED", x["pct"]))
    print(f"\n{'COLUMN':<56} {'FILL%':>6} {'BLANK':>6} {'ENC':>5}  STATUS")
    print("-" * 82)
    for r in issues:
        print(f"{r['col']:<56} {r['pct']:>5.1f}% {r['blank']:>6} {r['enc']:>5}  {r['status']}")

    print("\n=== Sample values — key columns ===")
    for kc in ["Occupation/Industry", "Collateral Description", "Collateral use",
               "Risk Categorization", "Constitution",
               "In case of self employed (SENP/ SEP)"]:
        matches = [h for h in headers if kc.lower() in h.lower()]
        if matches:
            idx = headers.index(matches[0])
            if idx < len(df.columns):
                top = df.iloc[:, idx].dropna().value_counts().head(8)
                print(f"\n  {matches[0][:65]}")
                for k, v in top.items():
                    print(f"    {str(k)[:68]!r}: {v}")

    filled_cols = sum(1 for i in range(len(df.columns)) if df.iloc[:, i].notna().any())
    log.info("Columns with data: %d / %d", filled_cols, len(df.columns))
    log.info("=" * 65)


def main() -> None:
    log.info("=" * 65)
    log.info("CUSTOM LAN SANCTION FORMAT EXTRACT  --  %s", TIMESTAMP)
    log.info("=" * 65)

    template_cols = get_template_columns()
    lans          = load_lans()

    cfg      = yaml.safe_load(CONFIG_PATH.read_text())["database"]
    conn_cfg = {k: cfg[k] for k in ("host", "database", "port", "user", "password")}

    log.info("Connecting to Redshift ...")
    conn = redshift_connector.connect(**conn_cfg, timeout=600)
    df   = run_query(conn, load_sql(lans))
    conn.close()

    write_excel(df, template_cols)
    quick_eval(df, template_cols)

    log.info("Output: %s", OUTPUT_PATH)
    try:
        os.startfile(str(OUTPUT_PATH))
    except Exception:
        pass


if __name__ == "__main__":
    main()
