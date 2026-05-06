"""
DA Pool Sanction Format – Bulk Extract  (~17K HE/LAP loans)
POS / Overdue / Rate from loan_dtl_cghfl = CURRENT_DATE live snapshot.
No {LAN_LIST} placeholder — universe is defined in SQL CTE target_lans.
"""
import yaml
import logging
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
import redshift_connector
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)

# ── Paths ──────────────────────────────────────────────────────────────────────
ROOT     = Path(__file__).resolve().parents[1]
ABFL_DIR = ROOT / "ABFL_Formats"
SQL_FILE = ABFL_DIR / "sql_da_pool_sanction_bulk.sql"
CFG_FILE = ROOT / "config" / "database.yaml"

TS          = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_PATH = ABFL_DIR / f"DA_Pool_sanction_format_{TS}.xlsx"
SHEET_NAME  = "Required for pool shortlisting"

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-6s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Columns genuinely not available in DB (orange headers) ────────────────────
NOT_FOUND = {
    "Subvention Scheme if any",
    "PMAY subsidy status \n(Claimed & received/ Claimed but not received)",
    "NRI Loan (Yes/No)",
    "ECLGS",
    "Link loan Part of pool (Yes/ No)",
    "Link Loan/Top up loan Number if applicable",
    "Legal Report ",
}

# ── Columns that are hardcoded / derived (blue headers) ──────────────────────
HARDCODED = {
    "PMAY Flag (Y/N) - All are HE cases thus marked as No",
    "EMI Frequency",
    "Current LTV at asset level",
    "financial applicant age",
    "Annual Income",
    "FOIR",
    "PDD status complete (Y/N)",
}

# ── Excel styles ──────────────────────────────────────────────────────────────
FILL_GREEN  = PatternFill("solid", fgColor="00B050")   # DB-sourced
FILL_ORANGE = PatternFill("solid", fgColor="FF6600")   # not in DB
FILL_BLUE   = PatternFill("solid", fgColor="0070C0")   # hardcoded/derived
FONT_WHITE  = Font(bold=True, color="FFFFFF", size=9)
FONT_DARK   = Font(bold=True, color="000000", size=9)

THIN  = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── DB connection ──────────────────────────────────────────────────────────────
def get_conn():
    cfg = yaml.safe_load(CFG_FILE.read_text())["database"]
    return redshift_connector.connect(
        host=cfg["host"],
        port=int(cfg.get("port", 5439)),
        database=cfg["database"],
        user=cfg["user"],
        password=cfg["password"],
        timeout=600,
    )


# ── SQL execution ──────────────────────────────────────────────────────────────
def run_query() -> pd.DataFrame:
    sql = SQL_FILE.read_text(encoding="utf-8").strip()

    logger.info("Connecting to Redshift …")
    conn = get_conn()
    try:
        cur = conn.cursor()
        logger.info("Executing bulk query (may take 2-5 min for full book) …")
        cur.execute(sql)
        cols = [d[0] for d in cur.description]
        logger.info("Fetching rows …")
        rows = cur.fetchall()
        df = pd.DataFrame(rows, columns=cols)
        logger.info("Fetched %d rows x %d columns", len(df), len(df.columns))
        return df
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ── Excel writer ───────────────────────────────────────────────────────────────
def write_excel(df: pd.DataFrame) -> None:
    logger.info("Writing Excel: %s", OUTPUT_PATH)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.freeze_panes = "C2"          # freeze row 1 + cols A-B

    # Header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
        if col_name in NOT_FOUND:
            cell.fill = FILL_ORANGE
            cell.font = FONT_WHITE
        elif col_name in HARDCODED:
            cell.fill = FILL_BLUE
            cell.font = FONT_DARK
        else:
            cell.fill = FILL_GREEN
            cell.font = FONT_DARK

    ws.row_dimensions[1].height = 60

    # Data rows
    logger.info("Writing data rows …")
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if isinstance(val, float) and str(val) == "nan":
                cell.value = None
            else:
                cell.value = val
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")

    # Auto column width (cap at 40)
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_data = df.iloc[:, col_idx - 1].astype(str)
        max_val  = col_data.str.len().max()
        header_w = len(str(col_name))
        if pd.notna(max_val):
            col_w = min(max(int(max_val), header_w, 10), 40)
        else:
            col_w = max(header_w, 12)
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = col_w

    wb.save(OUTPUT_PATH)
    logger.info("Saved: %s", OUTPUT_PATH)


# ── Evaluation summary ─────────────────────────────────────────────────────────
def eval_checks(df: pd.DataFrame) -> None:
    n = len(df)
    logger.info("── Eval summary ──────────────────────────────────────────")
    logger.info("Total rows: %d", n)

    filled = sum(1 for i in range(len(df.columns)) if df.iloc[:, i].notna().any())
    logger.info("  Columns with data : %d / %d", filled, len(df.columns))
    logger.info("  Columns not-in-DB : %d", len(NOT_FOUND))

    # Key column checks
    checks = [
        ("Loan No."             , lambda s: s.notna().sum() == n     , "All LANs populated"),
        ("Loan No."             , lambda s: ~s.duplicated()           , "No duplicate LANs"),
        ("Principal O/s"        , lambda s: pd.to_numeric(s, errors="coerce").fillna(0).ge(0).all(), "POS >= 0"),
        ("Current ROI"          , lambda s: pd.to_numeric(s, errors="coerce").dropna().between(1,40).all(), "ROI 1-40%"),
        ("FOIR"                 , lambda s: pd.to_numeric(s, errors="coerce").dropna().between(0,1).all(), "FOIR 0-1 (decimal)"),
        ("In case of self employed (SENP/ SEP)", lambda s: True, "SENP/SEP col present"),
    ]

    for col_name, fn, desc in checks:
        if col_name in df.columns:
            try:
                result = fn(df[col_name])
                if isinstance(result, pd.Series):
                    ok = result.all()
                else:
                    ok = bool(result)
                logger.info("  [%s] %s", "PASS" if ok else "FAIL", desc)
            except Exception as e:
                logger.info("  [SKIP] %s — %s", desc, e)

    # Profile type distribution
    if "Customer Sub-type (Primary Applicant)" in df.columns:
        dist = df["Customer Sub-type (Primary Applicant)"].value_counts()
        logger.info("  Borrower profile distribution:\n%s", dist.to_string())

    # SENP/SEP count
    senp_col = "In case of self employed (SENP/ SEP)"
    if senp_col in df.columns:
        senp_count = df[senp_col].notna().sum()
        logger.info("  SENP/SEP borrowers: %d / %d (%.1f%%)",
                    senp_count, n, senp_count / n * 100 if n else 0)

    logger.info("──────────────────────────────────────────────────────────")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    logger.info("=== DA Pool Sanction Format – Bulk Extract ===")
    logger.info("SQL  : %s", SQL_FILE)
    logger.info("Out  : %s", OUTPUT_PATH)
    logger.info("Date : %s (POS / DPD / Rate = CURRENT_DATE live from loan_dtl)", datetime.now().date())

    df = run_query()
    write_excel(df)
    eval_checks(df)

    logger.info("Done. Opening file …")
    try:
        os.startfile(str(OUTPUT_PATH))
    except Exception:
        pass


if __name__ == "__main__":
    main()
